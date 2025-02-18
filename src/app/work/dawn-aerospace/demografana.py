import sys
import random
import math
from openpyxl import Workbook
from datetime import datetime, timedelta

# Determine the correct date/time format based on OS (includes seconds)
if sys.platform.startswith("win"):
    date_format = "%#m/%#d/%y %#H:%M:%S"
else:
    date_format = "%-m/%-d/%y %-H:%M:%S"

def format_time(dt):
    return dt.strftime(date_format)

wb = Workbook()

# ------------------------------------------------------------------------------
# 1. LDR_ADC Data (1-second interval for 10 minutes)
# ------------------------------------------------------------------------------
ws = wb.active
ws.title = "LDR_ADC"
ws.append(["time"] + [f"LDR{i}" for i in range(1, 9)])

start = datetime(2025, 2, 18, 0, 0, 0)
total_seconds_ldr = 600  # 10 minutes
interval_ldr = 1         # 1 second per measurement
num_points_ldr = total_seconds_ldr // interval_ldr

for i in range(num_points_ldr):
    current_time = start + timedelta(seconds=i * interval_ldr)
    time_str = format_time(current_time)
    # Every 30-second cycle, a 5-second drop to ~600
    if (i % 30) < 5:
        ldr_values = [random.randint(580, 620) for _ in range(8)]
    else:
        ldr_values = [1023 for _ in range(8)]
    ws.append([time_str] + ldr_values)

# ------------------------------------------------------------------------------
# 2. Thruster Data (4 Thrusters, every 3 seconds for 5 minutes)
# ------------------------------------------------------------------------------
total_seconds_thruster = 300  # 5 minutes
interval_thruster = 3         # every 3 seconds
num_points_thruster = total_seconds_thruster // interval_thruster  # 100 points

for thruster in range(1, 5):
    ws_thruster = wb.create_sheet(title=f"Thruster_{thruster}")
    ws_thruster.append(["time", "vDIG", "vACT", "Current_Draw"])
    for i in range(num_points_thruster):
        current_time = start + timedelta(seconds=i * interval_thruster)
        time_str = format_time(current_time)

        if i % 10 == 0:
            # Spike every 30 seconds
            vDIG = 10.0
            vACT = 30.0
            current = 2.0
        else:
            # Baseline/incremental data
            vDIG = 12.0 + thruster + i * 0.05
            vACT = 11.5 + thruster + i * 0.04
            current = 5.0 + thruster * 0.1 + i * 0.02

        ws_thruster.append([time_str, round(vDIG, 2), round(vACT, 2), round(current, 2)])

# ------------------------------------------------------------------------------
# 3. Tank & Feedline Data (every second for 10 minutes, with 30-second heat cycling, offset baselines, and pressure)
# ------------------------------------------------------------------------------
ws_tf = wb.create_sheet(title="Tank_Feedline")
ws_tf.append(["time", "Tank_Ox", "Tank_FU", "Feedline_Ox", "Feedline_FU", "Tank_Pressure", "Feedline_Pressure"])

total_seconds_tf = 600  # 10 minutes
interval_tf = 1         # 1 second per measurement
num_points_tf = total_seconds_tf // interval_tf

# For a 30-second cycle, set cycle period accordingly.
cycle_period = 30  
# Define amplitudes (how much the cycle modulates the temperature)
amplitude_tank = 2.0      # Tanks have a smaller modulation amplitude
amplitude_feedline = 3.0  # Feedlines modulate more

for i in range(num_points_tf):
    current_time = start + timedelta(seconds=i * interval_tf)
    time_str = format_time(current_time)
    
    # Baseline: Tanks take longer to heat up (start lower, increase slower)
    tank_ox_baseline = 40 + i * 0.015
    tank_fu_baseline = 38 + i * 0.015
    # Feedlines heat up faster (start higher, increase quicker)
    feedline_ox_baseline = 45 + i * 0.02
    feedline_fu_baseline = 43 + i * 0.02
    
    # Sinusoidal modulation for heat cycling (30-second period)
    cycle = math.sin(2 * math.pi * (i / cycle_period))
    
    tank_ox = tank_ox_baseline + amplitude_tank * cycle
    tank_fu = tank_fu_baseline + amplitude_tank * cycle
    feedline_ox = feedline_ox_baseline + amplitude_feedline * cycle
    feedline_fu = feedline_fu_baseline + amplitude_feedline * cycle
    
    # Pressure data: simulate with a baseline and some random noise
    tank_pressure = 101 + random.uniform(-1, 1)       # e.g., around 101 kPa
    feedline_pressure = 100 + random.uniform(-1, 1)     # e.g., around 100 kPa

    ws_tf.append([
        time_str,
        round(tank_ox, 2),
        round(tank_fu, 2),
        round(feedline_ox, 2),
        round(feedline_fu, 2),
        round(tank_pressure, 2),
        round(feedline_pressure, 2)
    ])

# ------------------------------------------------------------------------------
# Save the workbook
# ------------------------------------------------------------------------------
output_filename = "dashboard_mock_data.xlsx"
wb.save(output_filename)
print(f"Excel file '{output_filename}' created successfully.")
